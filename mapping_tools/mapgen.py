#!/usr/bin/env python3

import sys
import os
import argparse
import json
import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from collections import OrderedDict


def noprint(*args, **kwargs):
	pass 

verbmsg = noprint
infomsg = print


DEVPARAMS = {
	'name'        : 'C1',
	'hwrev_from'  : 'C2',
	'hwrev_until' : 'C3',
	'map_type'    : 'C4',
	'map_rev'     : 'C5',
}

DEVFIRSTROW=8
DEVCOLLUT = {
	'offset'   : 2,
	'name'     : 3,
	'version'  : 4,
	'instance' : 5,
	'id'       : 6,
	'usize'    : 7,
	'size'     : 8,
	'mode'     : 9,
}

SECDEFFIRSTROW=2
SECDEFCOLLUT = {
	'id'       : 1,
	'name'   : 2,
}

SECFIRSTROW=2
SECCOLLUT = {
	'name'    : 1,
	'type'    : 2,
	'size'    : 3,
	'offset'  : 4,
	'mode'    : 5,
	'factor'  : 6,
	'exp'     : 7,
	'unit'    : 8,
	'min'     : 9,
	'max'     : 10,
	'default' : 11,
	'desc'    : 12,
	'info'    : 13,
}

MODELUT = {
	'default' : 'default',
	'no'      : 'no',
	'ro'      : 'R',
	'wo'      : ' W',
	'rw'      : 'RW',
}


if __name__ == '__main__':
	retcode = 0

	# Args parser config
	parser = argparse.ArgumentParser()
	parser.add_argument('sectiondef', help='XLSX sections definition input')
	parser.add_argument('devicedef', help='XLSX devices definition input')
	parser.add_argument('output', nargs='?', help='Output JSON file')
	parser.add_argument('-d', '--device', help='Device map to generate')
	parser.add_argument('-l', '--list', help='Just list the defined section and devices', action='store_true')
	parser.add_argument('-i', '--info', help='Show the device mapping', action='store_true')
	parser.add_argument('-v', '--verbose', help='Print more info', action='store_true')
	parser.add_argument('-q', '--quiet', help='Don\'t print anything', action='store_true')
	args = parser.parse_args()

	# Parse args
	if args.verbose:
		verbmsg = print
		infomsg = print
	if args.quiet:
		verbmsg = noprint
		infomsg = noprint
	verbmsg('args =', args)

	if not args.list and not args.device:
		infomsg('ERROR: must specify a device')
		sys.exit(1)

	# Open XLSX
	try:
		book_sec = load_workbook(filename=args.sectiondef, data_only=True)
		book_dev = load_workbook(filename=args.devicedef, data_only=True)
	except:
		infomsg('ERROR: fail to open workbook')
		sys.exit(1)

	try:
		sheet_sec_def = book_sec['__SectionDef']
	except:
		infomsg('ERROR: fail to access \'__SectionDef\' worksheet')
		sys.exit(1)

	# List devices
	devlist = []
	for d in book_dev.worksheets:
		if not d.title.startswith('!!__') and not d.title.startswith('$__') and not d.title.startswith('__REMOTE'):
			devlist.append(d.title)
	# end for d

	# Asked to print section list and device list
	if args.list:
		infomsg('============')
		infomsg('SECTION LIST')
		infomsg('')
		infomsg(f'{"ID   (dec)":<14} {"Name":<24} {"# Version"}')
		infomsg('--------------------------------------------------')
		sec_def_range = f'A{sheet_sec_def.min_row}:A{sheet_sec_def.max_row}'
		for row in sheet_sec_def.iter_rows(min_row=SECDEFFIRSTROW, values_only=True):
			if not row[0]:
				break
			# Minus 1 because LUT is in excel indexing
			secid = row[SECDEFCOLLUT['id']-1]
			secname = row[SECDEFCOLLUT['name']-1]
			# Count nbr version of the section
			secver = 0
			for s in book_sec.worksheets:
				if s.title.startswith(secname):
					secver = secver + 1
			infomsg(f'{f"0x{secid:02X} ({secid})":<14} {secname:<24} {secver}')
		# end for row

		infomsg('')
		infomsg('===========')
		infomsg('DEVICE LIST')
		infomsg('')
		for d in devlist:
			print(d)
		print('')
		sys.exit(0)


	# Reach this point -> generate json
	if args.device not in devlist:
		infomsg(f'ERROR: device {args.device} not found')
		sys.exit(1)

	jsout = {}
	jssecs = []
	outfile = args.output

	sheet_dev = book_dev[args.device]
	devname    = sheet_dev[DEVPARAMS['name']].value
	devhwfrom  = sheet_dev[DEVPARAMS['hwrev_from']].value
	devhwuntil = sheet_dev[DEVPARAMS['hwrev_until']].value
	devmaptype = sheet_dev[DEVPARAMS['map_type']].value
	devmaprev  = sheet_dev[DEVPARAMS['map_rev']].value

	# Compute the output file name if not provided
	#defoutfile = f'{devname.lower().replace(" ","_")}-{devhwfrom.lower()}-{devmaptype}_{devmaprev}_{datetime.datetime.now().strftime("%Y%m%d%H%M%S")}.json'
	defoutfile = f'{devname.lower().replace(" ","_")}-{devhwfrom.lower()}-{devmaptype}_{devmaprev}.json'
	if not outfile:
		outfile = defoutfile

	# Count the not special sections (not __InfoHeader, nor __SectionList)
	seccount = 0
	totalseccount = 0
	for drow in sheet_dev.iter_rows(min_row=DEVFIRSTROW, values_only=True):
		if drow[0] is None:
			break
		totalseccount = totalseccount + 1
		secname  = drow[DEVCOLLUT['name']-1]
		if (not secname.startswith('__InfoHeader')) and (not secname.startswith('__SectionList')):
			seccount = seccount + 1
	# end for drow

	devmsg = verbmsg
	if args.info:
		devmsg = infomsg
		infomsg('===========')
		infomsg('DEVICE INFO')
	else:
		infomsg(f'Generate mapping \'{outfile}\' from device definition \'{args.device}\'')

	devmsg('')
	devmsg(f'{args.device}')
	devmsg(f'> {"Name":<20} = {devname.lower().replace(" ","_")}')
	devmsg(f'> {"Default Output":<20} = {defoutfile}')
	devmsg(f'> {"HW rev":<20} = {devhwfrom} - {devhwuntil}')
	devmsg(f'> {"MB map":<20} = {devmaptype} - {devmaprev}')
	devmsg(f'> {"Section":<20} = {seccount}/{totalseccount}')
	devmsg('')
	devmsg(f'   {"MBADDR":>5}    {"NAME":<35}    {"MODE":<4}    {"EXTRA"}')
	devmsg('------------------------------------------------------------------------------------------')

	# Iter on each section
	for drow in sheet_dev.iter_rows(min_row=DEVFIRSTROW, values_only=True):
		jsregs = []
		if drow[0] is None:
			break
		secfound = False
		# Minus 1 because LUT is in excel indexing
		secoff   = drow[DEVCOLLUT['offset']-1]
		secname  = drow[DEVCOLLUT['name']-1]
		secrev   = drow[DEVCOLLUT['version']-1]
		secinst  = drow[DEVCOLLUT['instance']-1]
		secid    = drow[DEVCOLLUT['id']-1]
		secusize = drow[DEVCOLLUT['usize']-1]
		secsize  = drow[DEVCOLLUT['size']-1]
		secmode  = drow[DEVCOLLUT['mode']-1]
		secfullname = f'{secname}_v{secrev}'
		if secmode == 'default':
			tmpmode = 'ro'
		else:
			tmpmode = secmode

		devmsg('===')
		devmsg(f'SEC  {secoff:>5}    {f"{secfullname} (0x{secid:02X},0x{secrev:02X})":<35}    {f"{MODELUT[tmpmode]}":<4}    size={secinst}*{secusize}={secsize}')

		if secname == '__SectionList':
			if secinst != seccount:
				retcode = 10
				infomsg(f'WARNING: __SectionList instance should be {seccount} ({secinst} instead)')

		for s in book_sec.worksheets:
			if s.title == secfullname:
				secfound = True
		if not secfound:
			print(f'ERROR: section {secfullname} not found')
			sys.exit(1)

		sheet_sec = book_sec[secfullname]
		for srow in sheet_sec.iter_rows(min_row=SECFIRSTROW, values_only=True):
			if srow[0] is None:
				break
			regname = srow[SECCOLLUT['name']-1]
			regtype = srow[SECCOLLUT['type']-1]
			regsize = srow[SECCOLLUT['size']-1]
			regoff  = srow[SECCOLLUT['offset']-1]
			regmode = srow[SECCOLLUT['mode']-1]
			regfact = srow[SECCOLLUT['factor']-1]
			regexp  = srow[SECCOLLUT['exp']-1]
			regunit = srow[SECCOLLUT['unit']-1]
			regmin  = srow[SECCOLLUT['min']-1]
			regmax  = srow[SECCOLLUT['max']-1]
			regdef  = srow[SECCOLLUT['default']-1]
			regdesc = srow[SECCOLLUT['desc']-1]
			reginfo = srow[SECCOLLUT['info']-1]
			regaddr = secoff + regoff

			if regmode and (regmode != secmode) and (regmode != 'default'):
				tmpmode = regmode
			else:
				tmpmode = secmode

			jsreg = {
				'name'    : regname,
				'type'    : regtype,
				'size'    : regsize,
				'offset'  : regoff,
				}
			if regmode and (regmode != secmode) and (regmode != 'default'):
				jsreg['mode'] = regmode

			if regfact:
				jsreg['factor'] = regfact
			else:
				regfact = 1

			if regexp:
				jsreg['exp'] = regexp
			else:
				regexp = 0

			if regunit:
				jsreg['unit'] = regunit
			else:
				regunit = ''

			if regmin:
				jsreg['min'] = regmin

			if regmax:
				jsreg['max'] = regmax

			if regdef:
				jsreg['default'] = regdef

			if regdesc:
				jsreg['description'] = regdesc

			if reginfo:
				jsreg['$info'] = reginfo

			devmsg(f'REG  {regaddr:>5}    {regname:<35}    {f"{MODELUT[tmpmode]}":<4}    size={regsize:<3}  unit={regfact*(10**regexp)} {regunit}')

			jsregs.append(jsreg)
		# end for srow
		if secinst > 1:
			devmsg(f'... repeat x{secinst-1}')

		jssec = {
			'id'  : secid,
			'version'  : secrev,
			'name'     : secname,
			'address'  : secoff,
			'instance' : secinst,
			'isize'    : secusize,
			'size'     : secsize,
			'mode'     : secmode,
			'regs'     : jsregs
		}
		jssecs.append(jssec)
	# endfor drow

	jsout['sections'] = jssecs

	if not outfile:
		infomsg('ERROR: no output defined')
		sys.exit(1)
	if outfile == '-':
		print(json.dumps(jsout, indent='\t'))
	else:
		with open(outfile, 'w') as fdo:
			json.dump(jsout, fdo, indent='\t')

	sys.exit(retcode)
# end __main__



